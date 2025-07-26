import io
import traceback
from typing import Dict
import datetime # Added for dynamic filename

# Third-party imports
import pandas as pd
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# --- FastAPI App Initialization ---
app = FastAPI(
    title="Combined Excel Report Generator",
    description="Upload an Excel file to generate a multi-sheet report with Region Summary, Percentage Analysis, and KPI breakdowns.",
    version="1.0.0"
)

# CORS (Cross-Origin Resource Sharing) middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

# --- Helper Functions for Styling and Calculations ---

def get_cell_styles() -> Dict[str, object]:
    """
    Returns a dictionary of predefined openpyxl cell styles.
    Yeh function formatting ke liye styles (jaise border, font, color) define karta hai.
    """
    thin_side = Side(border_style="thin", color="000000")
    return {
        "border": Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side),
        "header_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "grand_total_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "red_fill": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "header_font": Font(bold=True, size=12),
        "bold_font": Font(bold=True),
        "centered_alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "data_font_red_bold": Font(bold=True, color="FF0000"), # New style for data in KPI sheets
        "total_font_bold": Font(bold=True),
        "total_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    }

def auto_fit_columns(ws, padding=4):
    """
    Auto-fits column widths based on content.
    Yeh function column ki width ko content ke hisab se adjust karta hai.
    """
    for col_cells in ws.columns:
        max_len = 0
        column = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if '\n' in str(cell.value): # Account for multiline headers
                        cell_len = max(len(line) for line in str(cell.value).split('\n'))
                    if cell_len > max_len:
                        max_len = cell_len
            except:
                pass
        adjusted_width = max_len + padding
        ws.column_dimensions[column].width = adjusted_width

# --- Core Logic from regionsheet.py ---

def create_region_summary_sheet(wb: Workbook, df: pd.DataFrame, filename: str):
    """
    Generates the 'Region Summary' sheet and adds it to the workbook.
    Yeh function 'regionsheet.py' ka logic istemal karke 'Region Summary' sheet banata hai.
    """
    try:
        ws = wb.create_sheet('Region Summary')
        styles = get_cell_styles()

        # Data processing
        df_region = df.copy()
        df_region.columns = df_region.columns.str.strip()
        df_region['REGION_NAME'] = df_region['REGION_NAME'].str.strip().str.upper()
        df_region['CO ORDINATOR NAME'] = df_region['CO ORDINATOR NAME'].fillna('').astype(str).str.strip()
        
        # Add STATE_NAME column and process it
        if 'STATE_NAME' in df_region.columns:
            df_region['STATE_NAME'] = df_region['STATE_NAME'].fillna('').astype(str).str.strip().str.upper()
        else:
            df_region['STATE_NAME'] = '' # Add an empty column if not present

        numeric_cols = [
            'TOTAL_FIN_SUCCESS', 'TOTAL_FIN_SUCCESS_AMT', 'TOTAL EKYC SUCCESS',
            'TOTAL APY SUCCESS', 'TOTAL PMSBY SUCCESS', 'TOTAL PMJJBY SUCCESS',
            'TOTAL LOAN RECOVERY', 'TOTAL AMOUNT', 'LOAN LEAD GENERATION COUNT'
        ]
        for col in numeric_cols:
            if col in df_region.columns:
                df_region[col] = pd.to_numeric(df_region[col], errors='coerce').fillna(0)

        # Include STATE_NAME in groupby and aggregation
        # Changed the groupby key to include 'STATE_NAME'
        grouped = df_region.groupby(['REGION_NAME', 'STATE_NAME']).agg(
            COORDINATOR_NAME=('CO ORDINATOR NAME', lambda x: '/'.join(sorted({i.strip().split()[0].upper() for i in x if i.strip()}))),
            TOTAL_BC=('REGION_NAME', 'count'),
            TOTAL_TXN_COUNT=('TOTAL_FIN_SUCCESS', 'sum'),
            TOTAL_TXN_AMOUNT=('TOTAL_FIN_SUCCESS_AMT', 'sum'),
            TOTAL_EKYC_SUCCESS=('TOTAL EKYC SUCCESS', 'sum'),
            TOTAL_APY_SUCCESS=('TOTAL APY SUCCESS', 'sum'),
            TOTAL_PMSBY_SUCCESS=('TOTAL PMSBY SUCCESS', 'sum'),
            TOTAL_PMJJBY_SUCCESS=('TOTAL PMJJBY SUCCESS', 'sum'),
            TOTAL_LOAN_RECOVERY=('TOTAL LOAN RECOVERY', 'sum'),
            TOTAL_RECOVERY_AMT=('TOTAL AMOUNT', 'sum'),
            TOTAL_LOAN_LEADS=('LOAN LEAD GENERATION COUNT', 'sum')
        ).reset_index()

        grouped = grouped.sort_values(by='TOTAL_TXN_COUNT', ascending=False)

        # Adjust total_row for 'STATE_NAME'
        total_row = {
            'REGION_NAME': 'GRAND TOTAL',
            'STATE_NAME': '', # STATE_NAME will be empty for grand total
            'COORDINATOR_NAME': '',
            'TOTAL_BC': grouped['TOTAL_BC'].sum(),
            'TOTAL_TXN_COUNT': grouped['TOTAL_TXN_COUNT'].sum(),
            'TOTAL_TXN_AMOUNT': grouped['TOTAL_TXN_AMOUNT'].sum(),
            'TOTAL_EKYC_SUCCESS': grouped['TOTAL_EKYC_SUCCESS'].sum(),
            'TOTAL_APY_SUCCESS': grouped['TOTAL_APY_SUCCESS'].sum(),
            'TOTAL_PMSBY_SUCCESS': grouped['TOTAL_PMSBY_SUCCESS'].sum(),
            'TOTAL_PMJJBY_SUCCESS': grouped['TOTAL_PMJJBY_SUCCESS'].sum(),
            'TOTAL_LOAN_RECOVERY': grouped['TOTAL_LOAN_RECOVERY'].sum(),
            'TOTAL_RECOVERY_AMT': grouped['TOTAL_RECOVERY_AMT'].sum(),
            'TOTAL_LOAN_LEADS': grouped['TOTAL_LOAN_LEADS'].sum(),
        }
        grouped = pd.concat([grouped, pd.DataFrame([total_row])], ignore_index=True)

        int_columns = ['TOTAL_BC', 'TOTAL_TXN_COUNT', 'TOTAL_EKYC_SUCCESS', 'TOTAL_APY_SUCCESS',
                       'TOTAL_PMSBY_SUCCESS', 'TOTAL_PMJJBY_SUCCESS', 'TOTAL_LOAN_RECOVERY', 'TOTAL_LOAN_LEADS']
        for col in int_columns:
            grouped[col] = grouped[col].astype(int)

        # Add STATE_NAME to column_rename
        column_rename = {
            'REGION_NAME': 'REGION', 'STATE_NAME': 'STATE', 'COORDINATOR_NAME': 'COORDINATOR', 'TOTAL_BC': 'TOTAL\nBC',
            'TOTAL_TXN_COUNT': 'TOTAL\nTXN\nCOUNT', 'TOTAL_TXN_AMOUNT': 'TOTAL\nTXN\nAMOUNT',
            'TOTAL_EKYC_SUCCESS': 'TOTAL\nEKYC', 'TOTAL_APY_SUCCESS': 'TOTAL\nAPY',
            'TOTAL_PMSBY_SUCCESS': 'TOTAL\nPMSBY', 'TOTAL_PMJJBY_SUCCESS': 'TOTAL\nPMJJBY',
            'TOTAL_LOAN_RECOVERY': 'TOTAL\nLOAN\nRECOVERY', 'TOTAL_RECOVERY_AMT': 'TOTAL\nRECOVERY\nAMT',
            'TOTAL_LOAN_LEADS': 'TOTAL\nLOAN\nLEADS',
        }
        grouped.rename(columns=column_rename, inplace=True)

        # Reorder columns to place STATE after REGION
        desired_order = ['REGION', 'STATE', 'COORDINATOR', 'TOTAL\nBC', 'TOTAL\nTXN\nCOUNT', 'TOTAL\nTXN\nAMOUNT',
                         'TOTAL\nEKYC', 'TOTAL\nAPY', 'TOTAL\nPMSBY', 'TOTAL\nPMJJBY', 'TOTAL\nLOAN\nRECOVERY',
                         'TOTAL\nRECOVERY\nAMT', 'TOTAL\nLOAN\nLEADS']
        # Filter desired_order to include only columns actually present in 'grouped'
        grouped = grouped[[col for col in desired_order if col in grouped.columns]]


        # Excel Formatting
        start_row, start_col = 5, 3
        title = filename.upper()
        max_col = start_col + len(grouped.columns) - 1
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=max_col)
        title_cell = ws.cell(row=start_row-1, column=start_col, value=title)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = styles["centered_alignment"]
        title_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        title_cell.border = styles["border"]

        header_colors = ['FFC000', '92D050', '00B0F0', 'FF99CC', 'BFBFBF', 'F4B084',
                         'A9D08E', 'D9E1F2', 'FCE4D6', 'DDEBF7', 'E2EFDA', 'FFF2CC']

        for i, col_name in enumerate(grouped.columns):
            cell = ws.cell(row=start_row, column=start_col + i, value=col_name)
            cell.font = styles["bold_font"]
            cell.alignment = styles["centered_alignment"]
            cell.fill = PatternFill(start_color=header_colors[i % len(header_colors)], fill_type='solid')
            cell.border = styles["border"]

        for row_idx, row in enumerate(grouped.itertuples(index=False), start=start_row + 1):
            # Check for 'GRAND TOTAL' in the first column, which is now 'REGION'
            is_grand_total = (row[0] == 'GRAND TOTAL')
            for col_idx, value in enumerate(row):
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                cell = ws.cell(row=row_idx, column=start_col + col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0'
                cell.alignment = styles["centered_alignment"]
                cell.font = styles["bold_font"]
                cell.border = styles["border"]
                if is_grand_total:
                    cell.fill = styles["grand_total_fill"]
                # Only apply red fill to numeric columns after 'COORDINATOR' and 'STATE'
                elif col_idx >= 3 and isinstance(value, (int, float)) and value == 0:
                    cell.fill = styles["red_fill"]
        
        # Apply autofit after all data is written
        auto_fit_columns(ws)


    except Exception:
        print("Error in create_region_summary_sheet:")
        print(traceback.format_exc())

# --- Core Logic from persantage.py ---

def create_percentage_sheet(wb: Workbook, df: pd.DataFrame):
    """
    Generates the 'PERCENTAGE' sheet and adds it to the workbook.
    Yeh function 'persantage.py' ka logic istemal karke 'PERCENTAGE' sheet banata hai.
    """
    try:
        ws = wb.create_sheet("PERCENTAGE")
        styles = get_cell_styles()

        # Data processing
        df_perc = df.copy()
        df_perc.columns = df_perc.columns.str.upper().str.strip().str.replace(" ", "_")

        # Add STATE_NAME column
        if 'STATE_NAME' in df_perc.columns:
            df_perc['STATE_NAME'] = df_perc['STATE_NAME'].fillna('').astype(str).str.strip().str.upper()
        else:
            df_perc['STATE_NAME'] = '' # Add an empty column if not present

        # Include STATE_NAME in groupby
        grouped = df_perc.groupby(["CO_ORDINATOR_NAME", "STATE_NAME"]) # Added STATE_NAME here
        result = grouped.agg(
            TOTAL_BC=("MECHNAT_ID", "count"),
            TOTAL_ACTIVE_BC=("TOTAL_LOGGING_DAYS", lambda x: (x != 0).sum())
        )
        result["TOTAL_INACTIVE_BC"] = result["TOTAL_BC"] - result["TOTAL_ACTIVE_BC"]
        result["%INACTIVE"] = (result["TOTAL_INACTIVE_BC"] / result["TOTAL_BC"] * 100).round(2)
        result["%AVG_TOTAL_FIN_SUCCESS"] = grouped["TOTAL_FIN_SUCCESS"].mean().round(2)
        result.reset_index(inplace=True)

        active_df = df_perc[df_perc["TOTAL_LOGGING_DAYS"] != 0]
        
        below_50_df = active_df[(active_df["TOTAL_FIN_SUCCESS"] >= 1) & (active_df["TOTAL_FIN_SUCCESS"] <= 50)]
        # Adjust groupby for below_50_df
        below_50 = below_50_df.groupby(["CO_ORDINATOR_NAME", "STATE_NAME"])["MECHNAT_ID"].count()
        # Ensure mapping works with multi-index
        result["BELOW_50_COUNT"] = result.set_index(["CO_ORDINATOR_NAME", "STATE_NAME"]).index.map(below_50).fillna(0).astype(int)
        result["BELOW_50_%"] = (result["BELOW_50_COUNT"] / result["TOTAL_ACTIVE_BC"] * 100).round(2)

        below_100_df = active_df[(active_df["TOTAL_FIN_SUCCESS"] >= 51) & (active_df["TOTAL_FIN_SUCCESS"] <= 100)]
        # Adjust groupby for below_100_df
        below_100 = below_100_df.groupby(["CO_ORDINATOR_NAME", "STATE_NAME"])["MECHNAT_ID"].count()
        # Ensure mapping works with multi-index
        result["BELOW_100_COUNT"] = result.set_index(["CO_ORDINATOR_NAME", "STATE_NAME"]).index.map(below_100).fillna(0).astype(int)
        result["BELOW_100_%"] = (result["BELOW_100_COUNT"] / result["TOTAL_ACTIVE_BC"] * 100).round(2)

        def calculate_kpi_percentage(data_frame, kpi_column):
            active_kpis = data_frame[data_frame["TOTAL_LOGGING_DAYS"] != 0]
            inactive_kpis = data_frame[data_frame["TOTAL_LOGGING_DAYS"] == 0]
            # Adjust groupby for active_kpi_count and inactive_kpi_count
            active_kpi_count = active_kpis.groupby(["CO_ORDINATOR_NAME", "STATE_NAME"])[kpi_column].apply(lambda x: (x > 0).sum())
            inactive_kpi_count = inactive_kpis.groupby(["CO_ORDINATOR_NAME", "STATE_NAME"])[kpi_column].apply(lambda x: (x > 0).sum())
            
            # Use multi-index for total_active_bc and total_inactive_bc
            total_active_bc = result.set_index(["CO_ORDINATOR_NAME", "STATE_NAME"])["TOTAL_ACTIVE_BC"]
            total_inactive_bc = result.set_index(["CO_ORDINATOR_NAME", "STATE_NAME"])["TOTAL_INACTIVE_BC"]
            
            active_percent = (active_kpi_count / total_active_bc * 100).round(2).fillna(0)
            inactive_percent = (inactive_kpi_count / total_inactive_bc * 100).round(2).fillna(0)
            return active_percent, inactive_percent

        kpi_columns = ["TOTAL_APY_SUCCESS", "TOTAL_PMSBY_SUCCESS", "TOTAL_PMJJBY_SUCCESS"]
        for kpi_col in kpi_columns:
            if kpi_col in df_perc.columns:
                active_pct, inactive_pct = calculate_kpi_percentage(df_perc, kpi_col)
                # Map back to original result DataFrame
                result[f"{kpi_col}_active_%"] = result.set_index(["CO_ORDINATOR_NAME", "STATE_NAME"]).index.map(active_pct).fillna(0)
                result[f"{kpi_col}_not_active_%"] = result.set_index(["CO_ORDINATOR_NAME", "STATE_NAME"]).index.map(inactive_pct).fillna(0)
        
        # Grand Total
        total_row = {"CO_ORDINATOR_NAME": "GRAND TOTAL", "STATE_NAME": ""} # Add STATE_NAME for grand total
        sum_cols = [col for col in result.columns if col not in ["CO_ORDINATOR_NAME", "STATE_NAME"] and not col.endswith("%")]
        for col in sum_cols:
            total_row[col] = result[col].sum()

        total_bc_grand = total_row.get("TOTAL_BC", 0)
        total_active_bc_grand = total_row.get("TOTAL_ACTIVE_BC", 0)
        total_inactive_bc_grand = total_row.get("TOTAL_INACTIVE_BC", 0)

        total_row["%INACTIVE"] = round(total_inactive_bc_grand / total_bc_grand * 100, 2) if total_bc_grand > 0 else 0
        total_row["%AVG_TOTAL_FIN_SUCCESS"] = round(active_df["TOTAL_FIN_SUCCESS"].mean(), 2) if not active_df.empty else 0
        total_row["BELOW_50_%"] = round(total_row.get("BELOW_50_COUNT", 0) / total_active_bc_grand * 100, 2) if total_active_bc_grand > 0 else 0
        total_row["BELOW_100_%"] = round(total_row.get("BELOW_100_COUNT", 0) / total_active_bc_grand * 100, 2) if total_active_bc_grand > 0 else 0

        for kpi_col in kpi_columns:
            if kpi_col in df_perc.columns:
                overall_active_kpi_count = df_perc[df_perc["TOTAL_LOGGING_DAYS"] != 0][kpi_col].gt(0).sum()
                overall_inactive_kpi_count = df_perc[df_perc["TOTAL_LOGGING_DAYS"] == 0][kpi_col].gt(0).sum()
                total_row[f"{kpi_col}_active_%"] = round((overall_active_kpi_count / total_active_bc_grand * 100), 2) if total_active_bc_grand > 0 else 0
                total_row[f"{kpi_col}_not_active_%"] = round((overall_inactive_kpi_count / total_inactive_bc_grand * 100), 2) if total_inactive_bc_grand > 0 else 0

        result = pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)

        column_rename_map = {
            "CO_ORDINATOR_NAME": "CO_ORDINATOR", "STATE_NAME": "STATE", "TOTAL_ACTIVE_BC": "ACTIVE_BC", "TOTAL_INACTIVE_BC": "INACTIVE_BC",
            "%AVG_TOTAL_FIN_SUCCESS": "%AVG_FIN_SUCCESS", "BELOW_50_COUNT": "BELOW_50_CNT", "BELOW_100_COUNT": "BELOW_100_CNT",
            "TOTAL_APY_SUCCESS_active_%": "APY_ACTIVE_%", "TOTAL_PMSBY_SUCCESS_active_%": "PMSBY_ACTIVE_%",
            "TOTAL_PMJJBY_SUCCESS_active_%": "PMJJBY_ACTIVE_%"
        }
        result.rename(columns=column_rename_map, inplace=True)

        final_cols = [
            "CO_ORDINATOR", "STATE", "TOTAL_BC", "ACTIVE_BC", "INACTIVE_BC", "%INACTIVE", "%AVG_FIN_SUCCESS",
            "BELOW_50_CNT", "BELOW_50_%", "BELOW_100_CNT", "BELOW_100_%", "APY_ACTIVE_%",
            "PMSBY_ACTIVE_%", "PMJJBY_ACTIVE_%"
        ]
        # Ensure all columns exist before selecting
        final_cols_exist = [col for col in final_cols if col in result.columns]
        result = result[final_cols_exist]

        # Excel writing
        last_row_index = len(result) + 1
        for r_idx, row in enumerate(dataframe_to_rows(result, index=False, header=True), start=1):
            ws.append(row)
            for cell in ws[r_idx]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = styles["border"]
                if r_idx == 1:
                    cell.font = styles["header_font"]
                    cell.fill = styles["header_fill"]
                elif r_idx == last_row_index:
                    cell.font = styles["bold_font"]
                    cell.fill = styles["grand_total_fill"]
                else:
                    cell.font = styles["bold_font"]

        auto_fit_columns(ws)

    except Exception:
        print("Error in create_percentage_sheet:")
        print(traceback.format_exc())

# --- Core Logic from sheet.py ---

def create_kpi_sheets(writer: pd.ExcelWriter, df: pd.DataFrame):
    """
    Generates 'Inactive', 'Below_50', 'Below_100', and 'SSS' sheets.
    Yeh function 'sheet.py' ka logic istemal karke bachi hui 4 sheets banata hai.
    """
    try:
        df_kpi = df.copy()
        df_kpi.columns = [str(c).strip() for c in df_kpi.columns]

        kpi_cols = [
            'TOTAL LOGGING DAYS', 'TOTAL EKYC SUCCESS', 'TOTAL APY SUCCESS',
            'TOTAL PMSBY SUCCESS', 'TOTAL PMJJBY SUCCESS', 'TOTAL LOAN RECOVERY',
            'TOTAL AMOUNT', 'LOAN LEAD GENERATION COUNT', 'TOTAL_FIN_SUCCESS'
        ]
        for col in kpi_cols:
            if col not in df_kpi.columns:
                df_kpi[col] = 0
            df_kpi[col] = pd.to_numeric(df_kpi[col], errors='coerce').fillna(0)
        
        # Add STATE_NAME to df_kpi and normalize
        if 'STATE_NAME' in df_kpi.columns:
            df_kpi['STATE_NAME'] = df_kpi['STATE_NAME'].fillna('').astype(str).str.strip().str.upper()
        else:
            df_kpi['STATE_NAME'] = '' # Add an empty column if not present


        base_columns = ['MECHNAT_ID', 'BC_NAME', 'BRANCH_NAME', 'REGION_NAME','STATE_NAME', 'LOCATION TYPE', 'TOTAL LOGGING DAYS', 'CO ORDINATOR NAME'] 
        below_columns = base_columns[:-1] + ['TOTAL_FIN_SUCCESS','CO ORDINATOR NAME']
        
        # Ensure base columns exist
        base_columns_exist = [col for col in base_columns if col in df_kpi.columns]
        below_columns_exist = [col for col in below_columns if col in df_kpi.columns]


        df_inactive = df_kpi[(df_kpi[kpi_cols].fillna(0) == 0).all(axis=1)].copy()[base_columns_exist]
        df_below_50 = df_kpi[(df_kpi['TOTAL_FIN_SUCCESS'] >= 1) & (df_kpi['TOTAL_FIN_SUCCESS'] <= 50)].copy()[below_columns_exist]
        df_below_100 = df_kpi[(df_kpi['TOTAL_FIN_SUCCESS'] >= 51) & (df_kpi['TOTAL_FIN_SUCCESS'] <= 100)].copy()[below_columns_exist]

        sss_filter = (
            (df_kpi['TOTAL LOGGING DAYS'] > 0) & (df_kpi['TOTAL APY SUCCESS'] == 0) &
            (df_kpi['TOTAL PMSBY SUCCESS'] == 0) & (df_kpi['TOTAL PMJJBY SUCCESS'] == 0)
        )
        sss_columns = [
            'MECHNAT_ID', 'BC_NAME', 'BRANCH_NAME', 'REGION_NAME','STATE_NAME', 'LOCATION TYPE',
            'TOTAL LOGGING DAYS', 'TOTAL APY SUCCESS', 'TOTAL PMSBY SUCCESS',
            'TOTAL PMJJBY SUCCESS', 'CO ORDINATOR NAME'
        ]
        sss_columns_exist = [col for col in sss_columns if col in df_kpi.columns]
        df_sss = df_kpi[sss_filter].copy()[sss_columns_exist]

        sheets_to_create = {
            'Inactive': df_inactive,
            'Below_50': df_below_50,
            'Below_100': df_below_100,
            'SSS': df_sss
        }

        # --- Writing to Excel using openpyxl engine with proper styling ---
        styles = get_cell_styles()

        for name, df_sheet in sheets_to_create.items():
            if df_sheet.empty: continue
            
            # Write the DataFrame to a new sheet
            df_sheet.to_excel(writer, sheet_name=name, index=False, startrow=1, header=False)
            ws = writer.sheets[name]
            
            # Apply header styling
            headers = ["\n".join(col.split()) for col in df_sheet.columns]
            for col_idx, header_text in enumerate(headers):
                cell = ws.cell(row=1, column=col_idx + 1, value=header_text)
                cell.font = styles["header_font"]
                cell.fill = styles["header_fill"]
                cell.alignment = styles["centered_alignment"]
                cell.border = styles["border"]
            ws.row_dimensions[1].height = 40 # Set header row height

            # Apply data styling
            for r_idx in range(len(df_sheet)):
                for c_idx, col in enumerate(df_sheet.columns):
                    cell = ws.cell(row=r_idx + 2, column=c_idx + 1, value=df_sheet.iloc[r_idx, c_idx])
                    cell.font = styles["data_font_red_bold"] # Apply red bold font for data
                    cell.alignment = styles["centered_alignment"]
                    cell.border = styles["border"]

            # Auto-fit columns using the helper function
            auto_fit_columns(ws)

            # Add Total Count row
            total_count_row_idx = len(df_sheet) + 2
            total_count_cell = ws.cell(row=total_count_row_idx, column=1, value="Total Count")
            total_count_cell.font = styles["total_font_bold"]
            total_count_cell.fill = styles["total_fill"]
            total_count_cell.alignment = styles["centered_alignment"]
            total_count_cell.border = styles["border"]

            count_value_cell = ws.cell(row=total_count_row_idx, column=2, value=len(df_sheet))
            count_value_cell.font = styles["total_font_bold"]
            count_value_cell.fill = styles["total_fill"]
            count_value_cell.alignment = styles["centered_alignment"]
            count_value_cell.border = styles["border"]


    except Exception:
        print("Error in create_kpi_sheets:")
        print(traceback.format_exc())

# --- FastAPI Endpoint ---

@app.post("/process-complete-report/", tags=["Excel Processing"])
async def process_complete_report(file: UploadFile = File(...)):

    # File extension check
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file (.xlsx, .xls).")

    try:
        # Read uploaded file content
        contents = await file.read()
        
        # Load the main DataFrame
        try:
            df = pd.read_excel(io.BytesIO(contents), sheet_name="DATA")
        except ValueError:
            raise HTTPException(status_code=400, detail="A sheet named 'DATA' was not found in the uploaded file.")

        # Create an in-memory BytesIO object to hold the final Excel file
        output_io = io.BytesIO()

        # --- Step 1: Create initial sheets with openpyxl ---
        # Load the original workbook to potentially preserve other sheets/macros
        wb = load_workbook(io.BytesIO(contents))
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb['Sheet'])

        create_region_summary_sheet(wb, df, file.filename)
        create_percentage_sheet(wb, df)
        
        # Save the workbook with the first two reports to our in-memory object
        wb.save(output_io)
        output_io.seek(0) # Rewind the buffer

        # --- Step 2: Append KPI sheets using pandas.ExcelWriter with 'openpyxl' engine ---
        with pd.ExcelWriter(output_io, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer: # Added if_sheet_exists='replace'
            # The writer will now append to the existing workbook in output_io
            create_kpi_sheets(writer, df)

        output_io.seek(0) # Rewind the buffer to the beginning for streaming

        # --- Step 3: Return the final file ---
        # Current date ko YYMMDD format mein lein
        current_date_str = datetime.datetime.now().strftime("%Y%m%d")
        
        # New filename banayein
        new_filename = f"Complete_Bank_Report_{current_date_str}.xlsx" 

        return StreamingResponse(
            output_io,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={new_filename}"}
        )

    except Exception as e:
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {e}")